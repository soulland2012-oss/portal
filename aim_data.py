"""Data access layer for the AIM Index (AI-maturity index) view.

Reads AIM_Index_SQB.xlsx — one row per department, 16 weighted metrics
(M1-M16, normalized 0-100) and a computed "Итоговый индекс" column.
Reads both 'AIM Index' (current month) and 'AIM Index Prev' (previous month)
for trend comparison.
"""
import os

import openpyxl

XLSX_FILE = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'AIM_Index_SQB.xlsx')
SHEET_NAME = 'AIM Index'
PREV_SHEET_NAME = 'AIM Index Prev'

HEADER_ROW = 5      # 1-indexed row with "№", "Подразделение", "M1".."M16", "AIM"
LABEL_ROW = 6        # metric display names
CATEGORY_ROW = 7     # metric category (Процессы, Данные, ...)
WEIGHT_ROW = 8       # metric weights (sum to 1)
FIRST_DATA_ROW = 9   # first department row


def _read_dept_values(ws, metrics_meta):
    """Return {dept_id: {'total': float, 'values': [v, ...]}} from a sheet."""
    result = {}
    for row in ws.iter_rows(min_row=FIRST_DATA_ROW):
        num = row[0].value
        name = row[1].value
        if num is None or name is None:
            continue
        values = [c.value for c in row[2:18]]
        metrics = [
            {'weight': meta['weight'], 'value': v if isinstance(v, (int, float)) else None}
            for meta, v in zip(metrics_meta, values)
        ]
        total = sum(m['weight'] * m['value'] for m in metrics if m['value'] is not None)
        result[int(num)] = {
            'total': round(total, 1),
            'values': [m['value'] for m in metrics],
        }
    return result


def build_aim():
    wb = openpyxl.load_workbook(XLSX_FILE, data_only=True)
    ws = wb[SHEET_NAME]

    labels = [c.value for c in ws[LABEL_ROW]][2:18]
    categories = [c.value for c in ws[CATEGORY_ROW]][2:18]
    weights = [c.value for c in ws[WEIGHT_ROW]][2:18]
    metric_codes = [c.value for c in ws[HEADER_ROW]][2:18]

    metrics_meta = [
        {'code': code, 'label': label, 'category': cat, 'weight': weight}
        for code, label, cat, weight in zip(metric_codes, labels, categories, weights)
    ]

    # Previous month data
    prev_map = {}
    if PREV_SHEET_NAME in wb.sheetnames:
        prev_map = _read_dept_values(wb[PREV_SHEET_NAME], metrics_meta)

    departments = []
    for row in ws.iter_rows(min_row=FIRST_DATA_ROW):
        num = row[0].value
        name = row[1].value
        if num is None or name is None:
            continue
        values = [c.value for c in row[2:18]]
        metrics = [
            {**meta, 'value': v if isinstance(v, (int, float)) else None}
            for meta, v in zip(metrics_meta, values)
        ]
        total = sum(m['weight'] * m['value'] for m in metrics if m['value'] is not None)

        prev = prev_map.get(int(num))
        prev_total = prev['total'] if prev else None
        if prev:
            for i, m in enumerate(metrics):
                m['prev_value'] = prev['values'][i]
        else:
            for m in metrics:
                m['prev_value'] = None

        departments.append({
            'id': int(num),
            'name': name.strip(),
            'total': round(total, 1),
            'prev_total': prev_total,
            'metrics': metrics,
        })

    return {'metrics_meta': metrics_meta, 'departments': departments}
